# data_storage.py
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataStorage:
    """Handle saving and loading measurement data from Excel files."""
    
    def __init__(self):
        self.data_file = Path.home() / "Documents" / "Data" / "FourPointMeasurement_Data.xlsx"
        self.data_file.parent.mkdir(parents=True, exist_ok=True)
        if OPENPYXL_AVAILABLE:
            self._ensure_workbook()
    
    def _ensure_workbook(self):
        """Create or verify the Excel workbook exists with proper headers."""
        if not OPENPYXL_AVAILABLE:
            return
        
        if not self.data_file.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Measurements"
            
            headers = [
                "Timestamp",
                "Thickness (mm)",
                "Wafer Area (cm²)",
                "Type",
                "Voltage (V)",
                "Current (A)",
                "Sheet Resistance (Ω/sq)",
                "Resistivity (Ω·cm)",
                "Conductivity (S/cm)",
                "Doping Concentration (cm⁻³)"
            ]
            
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="C3A66B", end_color="C3A66B", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions["A"].width = 20
            for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                ws.column_dimensions[col].width = 18
            
            wb.save(self.data_file)
    
    def save_measurement(self, thickness_mm: float, wafer_area_in: float, result: dict):
        """Save a measurement result to the Excel file."""
        if not OPENPYXL_AVAILABLE:
            return False
        
        try:
            # Ensure file exists before saving
            if not self.data_file.exists():
                self._ensure_workbook()
            
            wb = openpyxl.load_workbook(self.data_file)
            ws = wb.active
            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            row = [
                timestamp,
                thickness_mm,
                wafer_area_in,
                result["type"],
                result["voltage_v"],
                result["current_a"],
                result["sheet_res_ohm_per_sq"],
                result["resistivity_ohm_cm"],
                result["conductivity_s_per_cm"],
                result["doping_cm3"],
            ]
            
            ws.append(row)
            wb.save(self.data_file)
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            # Try to recover by recreating the file
            try:
                if self.data_file.exists():
                    os.remove(self.data_file)
                self._ensure_workbook()
                return self.save_measurement(thickness_mm, wafer_area_in, result)
            except Exception as recovery_error:
                print(f"Error recovering Excel file: {recovery_error}")
                return False
    
    def load_all_measurements(self) -> list:
        """Load all measurements from the Excel file."""
        if not OPENPYXL_AVAILABLE:
            return []
        
        # Ensure file exists
        if not self.data_file.exists():
            self._ensure_workbook()
            return []
        
        try:
            wb = openpyxl.load_workbook(self.data_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            measurements = []
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for i, header in enumerate(headers):
                    cell = row[i]
                    row_data[header] = cell.value
                measurements.append(row_data)
            
            return measurements
        except Exception as e:
            print(f"Error loading from Excel: {e}")
            # Try to recover by recreating the file
            try:
                if self.data_file.exists():
                    os.remove(self.data_file)
                self._ensure_workbook()
            except:
                pass
            return []
    
    def get_file_path(self) -> str:
        """Get the path to the Excel data file."""
        return str(self.data_file)
